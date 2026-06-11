"""
Email report generator for decoryedwards.com inboxes.
Reads from all three Gmail/Workspace inboxes and outputs an Excel report.

Usage:
    python fetch_emails.py
    python fetch_emails.py --days 30   # look back 30 days (default: 14)
    python fetch_emails.py --max 200   # max emails per inbox (default: 100)
"""

import argparse
import base64
import json
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

INBOXES = [
    "decory@decoryedwards.com",
    "tutoring@decoryedwards.com",
    "consulting@decoryedwards.com",
]

# Patterns used to classify emails by sender domain / subject keywords
TRANSACTIONAL_SENDERS = [
    "calendly.com",
    "stripe.com",
    "waveapps.com",
    "wave.com",
    "bluevine.com",
    "google.com",
    "googleworkspace.com",
    "no-reply",
    "noreply",
    "notifications@",
    "donotreply",
    "mailer",
    "automated",
    "formspree.io",
]

TUTORING_KEYWORDS = ["tutor", "lesson", "student", "homework", "math", "science",
                     "english", "test prep", "sat", "act", "school", "grade", "learn"]

CONSULTING_KEYWORDS = ["consult", "business", "strategy", "project", "proposal",
                       "contract", "services", "hire", "quote", "scope", "retainer"]

CALENDLY_SENDERS = ["calendly.com"]
STRIPE_SENDERS = ["stripe.com"]
WAVE_SENDERS = ["waveapps.com", "wave.com"]
BLUEVINE_SENDERS = ["bluevine.com"]

OUTPUT_DIR = Path(__file__).parent / "output"
CREDENTIALS_FILE = Path(__file__).parent / "credentials.json"
TOKEN_FILE = Path(__file__).parent / "token.json"
LAST_RUN_FILE = Path(__file__).parent / "last_run.json"


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def get_service():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                raise FileNotFoundError(
                    f"\ncredentials.json not found at {CREDENTIALS_FILE}\n"
                    "See README.md for setup instructions."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json())
    return build("gmail", "v1", credentials=creds)


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------

def fetch_messages(service, email: str, after_date: datetime, max_results: int) -> list[dict]:
    after_ts = int(after_date.timestamp())
    query = f"after:{after_ts}"
    rows = []
    page_token = None
    fetched = 0

    while fetched < max_results:
        batch = min(100, max_results - fetched)
        kwargs = dict(userId=email, q=query, maxResults=batch)
        if page_token:
            kwargs["pageToken"] = page_token

        result = service.users().messages().list(**kwargs).execute()
        messages = result.get("messages", [])
        if not messages:
            break

        for msg in messages:
            full = service.users().messages().get(
                userId=email, id=msg["id"], format="metadata",
                metadataHeaders=["From", "To", "Subject", "Date"]
            ).execute()
            rows.append(parse_message(full, email))
            fetched += 1
            if fetched >= max_results:
                break

        page_token = result.get("nextPageToken")
        if not page_token:
            break

    return rows


def parse_message(msg: dict, inbox: str) -> dict:
    headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
    sender = headers.get("From", "")
    subject = headers.get("Subject", "(no subject)")
    date_raw = headers.get("Date", "")

    # Parse date
    try:
        from email.utils import parsedate_to_datetime
        date_dt = parsedate_to_datetime(date_raw)
        date_str = date_dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        date_str = date_raw[:20] if date_raw else ""

    # Extract plain email address
    match = re.search(r"<(.+?)>", sender)
    sender_email = match.group(1).lower() if match else sender.lower().strip()
    sender_domain = sender_email.split("@")[-1] if "@" in sender_email else sender_email

    category = classify(sender_email, sender_domain, subject, inbox)
    status, action = recommended_action(category, subject)

    return {
        "Inbox": inbox,
        "Date": date_str,
        "Sender": sender_email,
        "Subject": subject,
        "Category": category,
        "Status": status,
        "Recommended Action": action,
    }


# ---------------------------------------------------------------------------
# Classify
# ---------------------------------------------------------------------------

def classify(sender_email: str, sender_domain: str, subject: str, inbox: str) -> str:
    subj_lower = subject.lower()

    # Check specific transactional services first
    for d in CALENDLY_SENDERS:
        if d in sender_domain:
            return "Calendly / Booking"
    for d in STRIPE_SENDERS:
        if d in sender_domain:
            return "Stripe / Payment"
    for d in WAVE_SENDERS:
        if d in sender_domain:
            return "Wave / Invoice"
    for d in BLUEVINE_SENDERS:
        if d in sender_domain:
            return "Bluevine / Banking"

    # Generic transactional
    for pattern in TRANSACTIONAL_SENDERS:
        if pattern in sender_email:
            return "Transactional / System"

    # Route by inbox first if keyword-ambiguous
    if inbox == "tutoring@decoryedwards.com":
        return "Tutoring Lead"
    if inbox == "consulting@decoryedwards.com":
        return "Consulting Lead"

    # Keyword-based for decory@ inbox
    if any(k in subj_lower for k in TUTORING_KEYWORDS):
        return "Tutoring Lead"
    if any(k in subj_lower for k in CONSULTING_KEYWORDS):
        return "Consulting Lead"

    return "General / Unknown"


def recommended_action(category: str, subject: str) -> tuple[str, str]:
    actions = {
        "Tutoring Lead":         ("New",    "Reply within 24h — confirm availability and send booking link"),
        "Consulting Lead":       ("New",    "Reply within 24h — schedule discovery call via Calendly"),
        "Calendly / Booking":    ("Info",   "Review booking details; add to calendar if not auto-synced"),
        "Stripe / Payment":      ("Info",   "Verify payment received; update Wave if needed"),
        "Wave / Invoice":        ("Info",   "Confirm invoice status; follow up if overdue"),
        "Bluevine / Banking":    ("Info",   "Review account alert"),
        "Transactional / System":("Info",   "No action needed unless unexpected"),
        "General / Unknown":     ("Review", "Read and categorize manually"),
    }
    status, action = actions.get(category, ("Review", "Categorize manually"))
    return status, action


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

CATEGORY_COLORS = {
    "Tutoring Lead":          "D6EAF8",  # blue
    "Consulting Lead":        "D5F5E3",  # green
    "Calendly / Booking":     "FEF9E7",  # yellow
    "Stripe / Payment":       "FDEDEC",  # red-light
    "Wave / Invoice":         "F4ECF7",  # purple-light
    "Bluevine / Banking":     "E8F8F5",  # teal
    "Transactional / System": "F2F3F4",  # grey
    "General / Unknown":      "FDFEFE",  # white
}

STATUS_COLORS = {
    "New":    "E74C3C",  # red text — needs action
    "Info":   "2E86C1",  # blue
    "Review": "F39C12",  # orange
}


def write_excel(rows: list[dict], output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Report"

    columns = ["Inbox", "Date", "Sender", "Subject", "Category", "Status", "Recommended Action"]
    col_widths = [32, 18, 36, 48, 24, 10, 52]

    # Header row
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (col, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        cat = row.get("Category", "")
        bg = CATEGORY_COLORS.get(cat, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg)

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == "Recommended Action"))

            if col == "Status":
                color = STATUS_COLORS.get(row.get("Status", ""), "000000")
                cell.font = Font(bold=True, color=color)

        ws.row_dimensions[row_idx].height = 18

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    from collections import Counter
    cat_counts = Counter(r["Category"] for r in rows)
    inbox_counts = Counter(r["Inbox"] for r in rows)

    ws2.append(["Category Breakdown"])
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.append(["Category", "Count"])
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        ws2.append([cat, count])

    ws2.append([])
    start_row = len(cat_counts) + 4
    ws2.cell(row=start_row, column=1, value="By Inbox").font = Font(bold=True, size=12)
    ws2.append(["Inbox", "Count"])
    for inbox, count in sorted(inbox_counts.items(), key=lambda x: -x[1]):
        ws2.append([inbox, count])

    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 10

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_last_run() -> datetime:
    if LAST_RUN_FILE.exists():
        data = json.loads(LAST_RUN_FILE.read_text())
        return datetime.fromisoformat(data["last_run"])
    # First ever run — default to 14 days back
    return datetime.now(timezone.utc) - timedelta(days=14)


def save_last_run(dt: datetime):
    LAST_RUN_FILE.write_text(json.dumps({"last_run": dt.isoformat()}))


def main():
    parser = argparse.ArgumentParser(description="Fetch and classify emails from decoryedwards.com inboxes")
    parser.add_argument("--max", type=int, default=100, help="Max emails per inbox (default: 100)")
    parser.add_argument("--reset", action="store_true", help="Ignore last-run timestamp and fetch last 14 days")
    args = parser.parse_args()

    run_start = datetime.now(timezone.utc)

    if args.reset or not LAST_RUN_FILE.exists():
        after_date = datetime.now(timezone.utc) - timedelta(days=14)
        print(f"Fetching emails from the last 14 days across {len(INBOXES)} inboxes...")
    else:
        after_date = load_last_run()
        print(f"Fetching emails since last run ({after_date.strftime('%Y-%m-%d %H:%M UTC')}) across {len(INBOXES)} inboxes...")

    service = get_service()
    all_rows = []

    for inbox in INBOXES:
        print(f"  Reading {inbox}...")
        try:
            rows = fetch_messages(service, inbox, after_date, args.max)
            print(f"    {len(rows)} emails found")
            all_rows.extend(rows)
        except Exception as e:
            print(f"    ERROR: {e}")

    # Sort by date descending
    all_rows.sort(key=lambda r: r["Date"], reverse=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = OUTPUT_DIR / f"email_report_{timestamp}.xlsx"
    write_excel(all_rows, output_path)

    save_last_run(run_start)

    print(f"\nDone. {len(all_rows)} total emails.")
    print(f"Report saved to: {output_path}")


if __name__ == "__main__":
    main()
